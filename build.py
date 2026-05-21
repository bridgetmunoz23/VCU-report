#!/usr/bin/env python3
"""
Build players.js for the VCU scouting report from raw data files.

Reads (per player <key>):
  kenpom_data/<key>_kenpom.txt   KenPom player snapshot (raw paste)
  play_types/<key>_plays.xlsx    Synergy play-type breakdown
  shot_types/<key>_shots.xlsx    Synergy shot-type breakdown
  shot_diet/<key>_diet.xlsx      Synergy per-shot event log (distance + FG flag)

Outputs:
  players.js   single `const players = {...}` object keyed by lowercase last
               name. Players auto-discovered from kenpom_data/.

The shotZones field is derived from the per-shot diet file: each row is one
shot with a distance ("23'") and a 2FGM/3FGm-style result code. Distance
banding: <=4 ft = restricted area, 4-14 ft = paint, 14+ ft = midrange.
Any 3FG (made or missed) goes to beyondArc regardless of distance.
"""

import json
import re
from pathlib import Path

import openpyxl


HERE = Path(__file__).parent
KENPOM_DIR = HERE / "kenpom_data"
PLAYS_DIR = HERE / "play_types"
SHOTS_DIR = HERE / "shot_types"
DIET_DIR = HERE / "shot_diet"
PICS_DIR = HERE / "photos"
OUT = HERE / "players.js"


# Per-player manual overrides. The archetype is a short pill string; analysis
# can be a string (paragraph) or list of strings (bulleted in the deep dive).
# Empty by default — the user writes these by hand once the data is in place.
PLAYER_OVERRIDES = {
    "lazar": {
        "archetype": "High-Volume Stretch 5",
        "analysis": [
            "34.3% from 3 on 67 attempts alongside 1.011 PPS Post-Up",
            "120.6 ORtg on 21.9% usage",
            "6.1 FD/40, 77.4 FTRate, 78.3% FT, elite at drawing fouls and converting",
        ],
    },
    "evans": {
        "archetype": "Transition 4",
        "analysis": [
            "31.7% from 3 on 82 attempts despite Spot Up being his top play type by volume",
            "1.194 PPS in Transition (Very Good) is where his offense lives, 0.816 PPS Spot Up (Below Average) is his shortcoming",
        ],
    },
    "hill": {
        "archetype": "Elite Three-Level Guard",
        "analysis": [
            "128.3 ORtg on 25.0% usage",
            "37.0% from 3 on 219 attempts and 85.7% from FT",
            "+2.3 FD/FC ratio, combined with his high FT% makes him both dangerous and comfortable at the line",
            "At least top 75th percentile in PPG for his 5 most frequent play types, ranging from .99 to 1.33 PPP",
        ],
    },
    "jennings": {
        "archetype": "Two-Way Combo Guard",
        "analysis": [
            "43.1% from 3 on 58 attempts",
            "3.5 Blk%, 3.9 Stl%, great defensively for a guard",
            "50.9 FTRate and 85.5% FT",
        ],
    },
    "lewis": {
        "archetype": "Standout Freshman Combo Guard",
        "analysis": [
            "118.8 ORtg on 18.4% usage as a freshman is well above average for his role",
            "36.2% from 3 on 127 attempts and 1.126 PPS Spot Up (Excellent), however numbers not competing with veteran shooters on roster",
            "+3.1 FD/FC ratio, with an average of 6 fouls per 40 mins",
        ],
    },
    "belle": {
        "archetype": "Efficient Stretch 4 / Elite Rebounder",
        "analysis": [
            "17.4 DR% and 7.4 OR% make him the team's best per-minute rebounder",
            "50.0% from 3 on 40 attempts (notably small sample but elite efficiency)",
        ],
    },
    "tracey": {
        "archetype": "High-Volume Wing",
        "analysis": [
            "Spot Up and P&R Ball Handler (1st and 3rd most frequent plays) his strongest, but ranks below average in transition his 2nd most frequent (17th percentile, 20% of plays)",
            "19.7 TORate is the highest among rotation players, ball security is the concern",
        ],
    },
    "ward": {
        "archetype": "Aggressive Wing",
        "analysis": [
            "57.5% on 2s vs 30.6% from 3 primarily shooting outside the rim",
            "43.0 FTRate paired with 21.8% usage, drives looking for contact",
            "1.444 PPS on Cuts (Excellent) on 27 poss, efficient off-ball mover",
        ],
    },
    "ahmad": {
        "archetype": "Spark Spot-Up Specialist",
        "analysis": [
            "11th percentile in transition",
            "41.4% from 3 on 58 attempts and 1.4 PPS Spot Up (99th percentile, his most frequent look)",
            "119.7 ORtg on 18.9% usage in limited minutes",
            "Small sample (20% min)",
        ],
    },
    "mitchell": {
        "archetype": "High-Energy 5 / Rebounder",
        "analysis": [
            "20.9 DR%",
            "1.471 PPS on Cuts (Excellent)",
            "20.8% from 3 and 61.8% FT, but must consider small minutes sample",
        ],
    },
}


# Play hierarchy: parent -> ordered list of valid sub-play names
PLAY_PARENTS = {
    "Spot Up":    ["No Dribble Jumper", "Drives Right", "Drives Left", "Drives Straight", "Turnover"],
    "Transition": ["Ballhandler", "Right Wing", "Left Wing", "Trailer", "Leak Outs", "First Middle"],
}
PLAY_TOP_LEVEL = {
    "Spot Up", "Transition", "P&R Ball Handler", "P&R Roll Man", "Isolation",
    "Post-Up", "Cut", "Handoffs", "Off Screen", "Miscellaneous Plays",
    "Offensive Rebounds (Put Backs)",
}

# Shot hierarchy: nested
SHOT_HIERARCHY = {
    "Jump Shot": {
        "Catch and Shoot": ["Guarded", "Unguarded"],
        "Dribble Jumper": [],
        "Early Jumper": [],
    },
    "At Rim": {
        "Layup": [],
        "Dunk": [],
        "Tip": [],
    },
    "Runner": {},
    "Hook": {},
}

YEAR_ABBR = {
    "Fr": "Freshman", "So": "Sophomore", "Jr": "Junior", "Sr": "Senior",
    "R-Fr": "R-Freshman", "R-So": "R-Sophomore", "R-Jr": "R-Junior", "R-Sr": "R-Senior",
    "5th": "5th Year",
}


# ---------- generic value parsing ----------

def to_float(cell):
    if cell is None:
        return None
    s = str(cell).strip()
    if s == "" or s == "-":
        return None
    s = s.replace("%", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def to_int(cell):
    f = to_float(cell)
    if f is None:
        return None
    return int(round(f))


def to_str(cell):
    if cell is None:
        return None
    s = str(cell).strip()
    if s == "" or s == "-":
        return None
    return s


# ---------- KenPom parser ----------

KENPOM_FIELDS = [
    "Ht", "Wt", "Yr", "G", "%Min", "ORtg", "%Poss", "%Shots", "eFG%", "TS%",
    "OR%", "DR%", "ARate", "TORate", "Blk%", "Stl%", "FC/40", "FD/40", "FTRate",
    "FTM-A", "FT_Pct", "2PM-A", "2P_Pct", "3PM-A", "3P_Pct",
]


def parse_kenpom(path):
    """Parse a KenPom paste. Tolerant of the several shapes the raw text can
    take across players (rank lines interspersed or not; team line on its own
    or merged with the first data row; with or without 'National Rank' header).
    """
    raw = Path(path).read_text(encoding="utf-8")
    lines = [l.rstrip("\n") for l in raw.split("\n") if l.strip() != ""]
    if len(lines) < 4:
        raise ValueError(f"{path}: too few non-empty lines ({len(lines)})")

    name = lines[0].strip()

    # Line 1: "<jersey> · <hometown>[ · <dob>]"
    bio_parts = [p.strip() for p in lines[1].split("·")]
    if len(bio_parts) < 2:
        raise ValueError(f"{path}: cannot parse bio line: {lines[1]!r}")
    jersey = bio_parts[0].lstrip("#").strip()
    try:
        jersey_num = int(jersey)
    except ValueError:
        jersey_num = jersey
    hometown = bio_parts[1]

    # Scan remaining lines for season, position, team, and data tokens
    season = None
    position = None
    team = ""
    data_tokens = []

    pos_re = re.compile(r"^[A-Z]{1,3}(·[A-Z]{1,3})*$")
    team_re = re.compile(r"^\d+\s+[A-Z][a-zA-Z]*")
    rank_re = re.compile(r"^-?\d+$")

    for line in lines[2:]:
        stripped = line.strip()

        # KenPom column-header row
        if "Ht" in stripped and "Wt" in stripped and "Yr" in stripped:
            continue
        # "National Rank" sub-header
        if stripped == "National Rank":
            continue
        # Season (4-digit year)
        if season is None and re.fullmatch(r"20\d\d", stripped):
            yr = int(stripped)
            season = f"{yr - 1}-{str(yr)[-2:]}"
            continue
        # Position
        if position is None and pos_re.fullmatch(stripped):
            position = stripped
            continue
        # Team line: "<rank> <team> <conf-rank> <region>" possibly followed by
        # tab-separated data tokens (raw paste sometimes merges them).
        if team == "" and team_re.match(line.split("\t")[0]):
            tabs = line.split("\t")
            head = tabs[0].split()
            if len(head) >= 2:
                team = head[1]
            for tok in tabs[1:]:
                if tok.strip():
                    data_tokens.append(tok.strip())
            continue
        # Lone integer = a national rank, skip
        if rank_re.fullmatch(stripped):
            continue
        # Otherwise: data line (tab-separated values)
        for tok in line.split("\t"):
            t = tok.strip()
            if t:
                data_tokens.append(t)

    if not season:
        raise ValueError(f"{path}: no season year found")

    # Pad if KenPom truncated
    if len(data_tokens) < len(KENPOM_FIELDS):
        data_tokens = data_tokens + [None] * (len(KENPOM_FIELDS) - len(data_tokens))

    raw_vals = dict(zip(KENPOM_FIELDS, data_tokens))

    def f(k):
        return to_float(raw_vals.get(k))

    def i(k):
        return to_int(raw_vals.get(k))

    def made_att(k):
        v = to_str(raw_vals.get(k))
        if not v or "-" not in v:
            return {"made": 0, "att": 0}
        try:
            made, att = v.split("-")
            return {"made": int(made), "att": int(att)}
        except ValueError:
            return {"made": 0, "att": 0}

    height = to_str(raw_vals.get("Ht")) or ""
    weight = i("Wt")
    year_abbr = to_str(raw_vals.get("Yr"))
    games = i("G")

    return {
        "name": name,
        "number": jersey_num,
        "position": position or "",
        "primary_position": (position or "").split("·")[0].strip(),
        "year_abbr": year_abbr,
        "year": YEAR_ABBR.get(year_abbr, year_abbr or ""),
        "height": height,
        "weight": weight,
        "hometown": hometown,
        "team": team or "VCU",
        "season": season,
        "games": games,
        "kenpom": {
            "minPct":   f("%Min"),
            "ortg":     f("ORtg"),
            "possPct":  f("%Poss"),
            "shotsPct": f("%Shots"),
            "efg":      f("eFG%"),
            "ts":       f("TS%"),
            "orPct":    f("OR%"),
            "drPct":    f("DR%"),
            "aRate":    f("ARate"),
            "toRate":   f("TORate"),
            "blkPct":   f("Blk%"),
            "stlPct":   f("Stl%"),
            "fc40":     f("FC/40"),
            "fd40":     f("FD/40"),
            "ftRate":   f("FTRate"),
            "ft":       {**made_att("FTM-A"), "pct": f("FT_Pct")},
            "twoP":     {**made_att("2PM-A"), "pct": f("2P_Pct")},
            "threeP":   {**made_att("3PM-A"), "pct": f("3P_Pct")},
        },
    }


# ---------- Synergy plays/shots parser ----------

def load_synergy(path):
    """Return (header_list, list_of_row_dicts)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{path}: empty sheet")
    header = rows[0]
    data = []
    for r in rows[1:]:
        d = dict(zip(header, r))
        if to_str(d.get("stat")):
            data.append(d)
    return header, data


def extract_row(d, eff_key):
    pps = to_float(d.get("pps"))
    if pps is None:
        pps = to_float(d.get(eff_key))
    poss = to_int(d.get("poss"))
    time_pct = to_float(d.get("time"))
    efg = to_float(d.get("efg"))
    if eff_key == "pps":
        rating = to_str(d.get("ppsrating"))
        rank = to_int(d.get("ppsrank"))
    else:
        rating = to_str(d.get("ppprating"))
        rank = to_int(d.get("ppprank"))
    return {
        "name": to_str(d.get("stat")),
        "poss": poss if poss is not None else 0,
        "timePct": time_pct,
        "pps": pps,
        "efg": efg,
        "verdict": rating,
        "percentile": rank,
    }


def parse_plays(path):
    _, rows = load_synergy(path)
    parsed = [extract_row(r, "ppp") for r in rows]

    by_name = {p["name"]: p for p in parsed}
    in_file = {p["name"] for p in parsed}
    result = []
    used = set()

    for p in parsed:
        nm = p["name"]
        if nm in used:
            continue
        if nm in PLAY_TOP_LEVEL:
            p = dict(p)
            p["subs"] = []
            if nm in PLAY_PARENTS:
                for sub_name in PLAY_PARENTS[nm]:
                    if sub_name in by_name and sub_name in in_file and sub_name != nm:
                        sub = by_name[sub_name]
                        p["subs"].append({
                            "name": sub_name,
                            "poss": sub["poss"],
                            "timePct": sub["timePct"],
                            "pps": sub["pps"],
                            "efg": sub["efg"],
                        })
                        used.add(sub_name)
            used.add(nm)
            result.append(p)

    total_top = sum(p["poss"] for p in result if p["poss"])
    for p in result:
        if p["timePct"] is None and total_top:
            p["timePct"] = round(100 * p["poss"] / total_top, 1)
        if p.get("subs"):
            parent_poss = p["poss"] or sum(s["poss"] for s in p["subs"])
            for s in p["subs"]:
                if s["timePct"] is None and parent_poss:
                    s["timePct"] = round(100 * s["poss"] / parent_poss, 1)

    return result


def parse_shots(path):
    _, rows = load_synergy(path)
    parsed = [extract_row(r, "pps") for r in rows]
    by_name = {p["name"]: p for p in parsed}
    in_file = {p["name"] for p in parsed}

    result = []
    for top_name, children in SHOT_HIERARCHY.items():
        if top_name not in in_file:
            continue
        top = dict(by_name[top_name])
        top["subs"] = []
        if isinstance(children, dict):
            for sub_name, sub_children in children.items():
                if sub_name not in in_file:
                    continue
                sub = by_name[sub_name]
                sub_entry = {
                    "name": sub_name,
                    "poss": sub["poss"],
                    "timePct": sub["timePct"],
                    "pps": sub["pps"],
                    "efg": sub["efg"],
                    "subs": [],
                }
                for grand_name in sub_children:
                    if grand_name in in_file:
                        g = by_name[grand_name]
                        sub_entry["subs"].append({
                            "name": grand_name,
                            "poss": g["poss"],
                            "timePct": g["timePct"],
                            "pps": g["pps"],
                            "efg": g["efg"],
                        })
                top["subs"].append(sub_entry)
        result.append(top)

    return result


# ---------- Shot zone derivation from per-shot diet log ----------

DIST_RE = re.compile(r"^(\d+)'$")
FG_RE = re.compile(r"^([23])FG([Mm])$")


def parse_diet_zones(path):
    """Walk the per-shot diet log and bucket attempts into 4 zones.

    The web-scraper output puts the distance and FG-result codes in
    different columns across files (we've seen 2FGM/3FGm in indexes
    10 or 11 depending on the export). Scan each row for tokens that
    match the distance and FG regexes rather than relying on a fixed
    column index.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    zones = {
        "restrictedArea": {"att": 0, "fgm": 0, "tpm": 0},
        "paint":          {"att": 0, "fgm": 0, "tpm": 0},
        "midrange":       {"att": 0, "fgm": 0, "tpm": 0},
        "beyondArc":      {"att": 0, "fgm": 0, "tpm": 0},
    }

    for row in rows[1:]:
        distance = None
        fg_made = None
        is_three = False
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if distance is None:
                m = DIST_RE.match(s)
                if m:
                    distance = int(m.group(1))
                    continue
            if fg_made is None:
                m = FG_RE.match(s)
                if m:
                    is_three = (m.group(1) == "3")
                    fg_made = (m.group(2) == "M")

        if fg_made is None:
            continue

        if is_three:
            zone = "beyondArc"
        elif distance is None:
            continue
        elif distance <= 4:
            zone = "restrictedArea"
        elif distance <= 14:
            zone = "paint"
        else:
            zone = "midrange"

        zones[zone]["att"] += 1
        if fg_made:
            zones[zone]["fgm"] += 1
            if is_three:
                zones[zone]["tpm"] += 1

    total_att = sum(z["att"] for z in zones.values())
    result = {}
    for key, c in zones.items():
        att = c["att"]
        vol = round(100 * att / total_att, 1) if total_att else 0
        efg = round(100 * (c["fgm"] + 0.5 * c["tpm"]) / att, 1) if att else None
        result[key] = {"volume": vol, "efg": efg, "poss": att}

    return result


# ---------- Build ----------

def discover_players():
    """Player keys = stems of <key>_kenpom.txt files, sorted alphabetically."""
    keys = []
    for p in sorted(KENPOM_DIR.glob("*_kenpom.txt")):
        stem = p.stem
        if stem.endswith("_kenpom"):
            keys.append(stem[: -len("_kenpom")])
    return keys


def build_player(key):
    kp_path = KENPOM_DIR / f"{key}_kenpom.txt"
    plays_path = PLAYS_DIR / f"{key}_plays.xlsx"
    shots_path = SHOTS_DIR / f"{key}_shots.xlsx"
    diet_path = DIET_DIR / f"{key}_diet.xlsx"
    photo_path = PICS_DIR / f"{key}.png"

    if not kp_path.exists():
        raise FileNotFoundError(f"[{key}] missing kenpom file: {kp_path}")

    bio = parse_kenpom(kp_path)
    plays = parse_plays(plays_path) if plays_path.exists() else []
    shots = parse_shots(shots_path) if shots_path.exists() else []
    zones = parse_diet_zones(diet_path) if diet_path.exists() else {
        "restrictedArea": {"volume": 0, "efg": None, "poss": 0},
        "paint":          {"volume": 0, "efg": None, "poss": 0},
        "midrange":       {"volume": 0, "efg": None, "poss": 0},
        "beyondArc":      {"volume": 0, "efg": None, "poss": 0},
    }

    overrides = PLAYER_OVERRIDES.get(key, {})
    return {
        **bio,
        "key": key,
        "photo": f"photos/{key}.png" if photo_path.exists() else None,
        "playTypes": plays,
        "shotTypes": shots,
        "shotZones": zones,
        "archetype": overrides.get("archetype", ""),
        "analysis":  overrides.get("analysis",  ""),
    }


def main():
    keys = discover_players()
    if not keys:
        raise SystemExit(f"No kenpom files found under {KENPOM_DIR}")

    built = {}
    errors = []
    for key in keys:
        try:
            built[key] = build_player(key)
            print(f"  built {key}: {built[key]['name']}")
        except Exception as e:
            errors.append(f"[{key}] {e}")
            print(f"  FAILED {key}: {e}")

    if not built:
        raise SystemExit("No players built — aborting.")

    # Sort by %Min descending so the consuming HTML can rely on iteration order
    ordered = dict(sorted(
        built.items(),
        key=lambda kv: (kv[1]["kenpom"].get("minPct") or 0),
        reverse=True,
    ))

    js = "// AUTO-GENERATED by build.py — do not edit by hand.\n"
    js += "// Re-run `python3 build.py` to regenerate.\n"
    js += "const players = " + json.dumps(ordered, indent=2, ensure_ascii=False) + ";\n"
    OUT.write_text(js, encoding="utf-8")
    print(f"\nWrote {OUT} ({len(ordered)} player(s))")

    if errors:
        print("\nErrors:")
        for e in errors:
            print("  " + e)


if __name__ == "__main__":
    main()
